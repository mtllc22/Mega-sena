from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import *
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from datetime import datetime
import time
import os
import csv


# =====================================================================
# PASTAS
# =====================================================================
LOG_DIR = "logs"
PRINT_DIR = "prints"
REL_DIR = "relatorios"

for d in [LOG_DIR, PRINT_DIR, REL_DIR]:
    if not os.path.exists(d):
        os.makedirs(d)

# =====================================================================
# ARQUIVOS
# =====================================================================
LOG_FILE = f"{LOG_DIR}/mega_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
RELATORIO = f"{REL_DIR}/mega_resultados_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.csv"


# =====================================================================
# LOG
# =====================================================================
def log(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{timestamp}] {msg}"
    print(linha)
    with open(LOG_FILE, "a") as f:
        f.write(linha + "\n")


# =====================================================================
# PRINT AUTOMÁTICO
# =====================================================================
def salvar_print(driver, titulo):
    nome = f"{PRINT_DIR}/{titulo}_{int(time.time())}.png"
    driver.save_screenshot(nome)
    log(f"📸 Print salvo: {nome}")
    return nome


# =====================================================================
# RELATÓRIO FINAL
# =====================================================================
with open(RELATORIO, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["Jogo Nº", "Dezenas", "Status", "Print", "Obs"])


def registrar_resultado(numero, dezenas, status, imagem, obs):
    with open(RELATORIO, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([numero, dezenas, status, imagem, obs])


# =====================================================================
# CONFIGURAÇÕES
# =====================================================================
PAUSA_A_CADA = 2
TEMPO_PAUSA = 2
MAX_TENTATIVAS_CLICK = 5

URL_HOME = "https://www.loteriasonline.caixa.gov.br"
URL_MEGA = "https://www.loteriasonline.caixa.gov.br/silce-web/#/mega-sena/especial"

JOGOS_ENVIADOS = set()   # <- DETECÇÃO DE REPETIDOS


# =====================================================================
# LEITURA DA PLANILHA
# =====================================================================
def ler_jogos():
    wb = load_workbook("mega.xlsx")
    ws = wb["mega"]
    jogos = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        dezenas = row[1:7]
        dezenas = [int(x) for x in dezenas if x is not None]
        if len(dezenas) == 6:
            jogos.append(tuple(sorted(dezenas)))   # <-- sorted para comparar

    return jogos


# =====================================================================
# ABERTURA DO SITE
# =====================================================================
def abrir_site():
    log("Abrindo navegador...")
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")

    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.get(URL_HOME)
    return driver


# =====================================================================
# POPUP SIM
# =====================================================================
def clicar_sim(driver):
    wait = WebDriverWait(driver, 10)
    try:
        sim_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//a[contains(text(),'Sim')]")
        ))
        driver.execute_script("arguments[0].click();", sim_btn)
        time.sleep(2)
        log("✔ Clique em SIM realizado.")
    except:
        log("ℹ Nenhum popup SIM encontrado.")


# =====================================================================
# FECHAR POPUPS
# =====================================================================
def fechar_popups(driver):
    xpaths = [
        "//button[contains(text(),'Aceitar')]",
        "//button[contains(text(),'Fechar')]",
        "//button[contains(text(),'OK')]",
        "//button[contains(text(),'Confirmar')]",
    ]

    for xp in xpaths:
        try:
            el = driver.find_element(By.XPATH, xp)
            driver.execute_script("arguments[0].click();", el)
            log(f"✔ Popup fechado: {xp}")
            time.sleep(1)
        except:
            pass


# =====================================================================
# LIMPAR VOLANTE
# =====================================================================
def limpar_volante(driver):
    try:
        wait = WebDriverWait(driver, 10)
        bot = wait.until(EC.presence_of_element_located((By.ID, "limparvolante")))
        driver.execute_script("arguments[0].click();", bot)
        time.sleep(0.4)
        log("🧽 Volante limpo com sucesso.")
        return True
    except Exception as e:
        log(f"⚠ Erro ao limpar volante: {e}")
        return False


# =====================================================================
# CLICAR DEZENA
# =====================================================================
def clicar_dezena(driver, dez):
    wait = WebDriverWait(driver, 10)
    num = f"{dez:02d}"
    element_id = f"n{num}"

    for tentativa in range(1, MAX_TENTATIVAS_CLICK + 1):
        try:
            el = wait.until(EC.presence_of_element_located((By.ID, element_id)))
            driver.execute_script("arguments[0].click();", el)
            time.sleep(0.25)

            if "selected" in el.get_attribute("class"):
                log(f"✔ Dezena {num} marcada!")
                return True

            log(f"❗ NÃO marcou {num} (tentativa {tentativa})")

        except Exception as e:
            log(f"⚠ Erro ao clicar {num}: {e}")

    log(f"❌ Falha definitiva ao marcar {num}")
    return False


# =====================================================================
# MARCAR JOGO
# =====================================================================
def marcar_jogo(driver, dezenas):
    log(f"📌 Lançando jogo: {dezenas}")

    for d in dezenas:
        if not clicar_dezena(driver, d):
            log("🚫 Falha ao marcar o jogo.")
            limpar_volante(driver)
            return False

    return True


# =====================================================================
# ENVIAR JOGO AO CARRINHO
# =====================================================================
def enviar_jogo(driver):
    wait = WebDriverWait(driver, 20)

    try:
        bot = wait.until(EC.presence_of_element_located((By.ID, "colocarnocarrinho")))
        driver.execute_script("arguments[0].click();", bot)
        log("✔ Jogo enviado para o carrinho.")
        time.sleep(1)
        return True

    except Exception as e:
        log(f"⚠ Erro ao enviar para o carrinho: {e}")
        log("🧽 Limpando volante...")
        limpar_volante(driver)
        return False


# =====================================================================
# PROCESSO PRINCIPAL
# =====================================================================
def principal():
    jogos = ler_jogos()
    total = len(jogos)

    log(f"🔎 {total} jogos encontrados na planilha mega.xlsx")

    driver = abrir_site()
    time.sleep(2)

    clicar_sim(driver)
    fechar_popups(driver)
    driver.get(URL_MEGA)
    time.sleep(3)

    for i, jogo in enumerate(jogos, start=1):
        dezenas = tuple(sorted(jogo))

        # 🚫 detectar repetidos
        if dezenas in JOGOS_ENVIADOS:
            log(f"⚠ Jogo repetido detectado e pulado: {dezenas}")
            registrar_resultado(i, dezenas, "REPETIDO", "", "Jogo duplicado")
            continue

        img_ini = salvar_print(driver, f"antes_jogo_{i}")

        if not marcar_jogo(driver, dezenas):
            log("🚫 Erro ao marcar dezenas")
            img = salvar_print(driver, f"erro_marcar_{i}")
            registrar_resultado(i, dezenas, "ERRO", img, "Falha ao marcar dezenas")
            continue

        img_mar = salvar_print(driver, f"marcado_jogo_{i}")

        if enviar_jogo(driver):
            img_env = salvar_print(driver, f"enviado_{i}")
            registrar_resultado(i, dezenas, "OK", img_env, "Enviado com sucesso")
            JOGOS_ENVIADOS.add(dezenas)

        else:
            img = salvar_print(driver, f"erro_envio_{i}")
            registrar_resultado(i, dezenas, "ERRO", img, "Falha ao enviar")
            continue

        if i % PAUSA_A_CADA == 0:
            log(f"⏸ Pausa automática de {TEMPO_PAUSA}s")
            time.sleep(TEMPO_PAUSA)

    log("🎯 FINALIZADO! Relatório criado em:")
    log(RELATORIO)
    return driver


# =====================================================================
# EXECUTAR
# =====================================================================
if __name__ == "__main__":
    principal()
